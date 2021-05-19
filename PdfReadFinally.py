import time

import pdfplumber
import numpy as np
import xlwt
from openpyxl import Workbook
from pdfplumber import table
#workbook = xlwt.Workbook()  #定义workbook
#sheet = workbook.add_sheet('Sheet1')  #添加sheet
now_time = time.strftime("%Y-%m-%d %H-%M-%S-", time.localtime())
with pdfplumber.open(r'D:\AS3000\PdfRead\20210513.pdf') as pdf:
    # 前提所有列必须固定
    table1 = np.array([
        #'电厂名称' ,'发电计划考核费用(万元)' ,'一次调频考核费用(万元)' ,'电压曲线合格率考核费用(万元)' ,'调峰考核费用(万元)',
        #'燃料管理考核费用(万元)', '技术监督考核费用(万元)', '安全管理考核费用(万元)' ,'调度运行管理考核费用（万元）',
        #'黑启动考核费用(万元)' ,'风电风功率预测考核费用(万元)' ,'风电有功功率变化考核(万元)', '风电脱网考核(万元)',
        #'风电发电计划考核费用(万元)', '调度检修管理考核费用(万元)', 'AVC考核费用(万元)' ,'不含非停考核费用合计(万元)',
        #'不含非停考核返还费用(万元)', '非停考核费用(万元)' ,'非停返还费用(万元)', '并网运行考核合计净收入(万元)' ,'上网电量(兆瓦时)'
        '电厂名称', '发电计划考核费用（万元）', '一次调频考核费用（万元）', 'AGC考核费用（万元）', '电压曲线合格率考核费用（万元）', '调峰考核费用（万元）', '燃料管理考核费用（万元）', '技术监督考核费用（万元）', '安全管理考核费用（万元）', '调度管理费（万元）', '黑启动考核费用（万元）', '风电风功率预测考核费用（万元）', '风电有功功率变化考核（万元）', '风电脱网考核（万元）', '风电发电计划考核费用（万元）', '调度检修管理考核费用（万元）', 'AVC考核费用（万元）', '不含非停考核费用合计（万元）', '不含非停考核返还费用（万元）', '非停考核费用（万元）', '非停返还费用（万元）', '并网运行考核合计净收入（万元）', '上网电量（MWh）'
    ])
    table2 = np.array([
        #'电厂名称', '黑启动补偿净收入(万元)', '冷备用补偿净收入(万元)', '无功补偿净收入(万元)', '旋转备用补偿净收入(万元)', '启停调峰补偿净收入(万元)', '深度调峰补偿净收入(万元)', '辅助服务补偿合计净收入(万元)'
        '电厂名称', 'AGC补偿净收入（万元）', 'AVC补偿净收入（万元）', '黑启动补偿净收入（万元）', '冷备用补偿净收入（万元）', '无功补偿净收入（万元）', '旋转备用补偿净收入（万元）', '启停调峰补偿净收入（万元）', '深度调峰补偿净收入（万元）', '辅助服务补偿合计净收入（万元）'
    ])
    table1List1 = []  # 表1数据集合
    table1List2 = []  # 表1数据集合
    #i = 0
    for page in pdf.pages:
        # page = pdf.pages[index]  # 页数 从0开始
        top1 = []
        index = 0
        flagtop = False
        for table in page.extract_tables():
            for row in table:
                topexsit = False
                if index == 0:
                    for cell in row:
                        if row[0].replace('\n', '').replace('\r', '') == "电厂名称":
                            top1.append("".join(cell.split()))  # 将第一行的集合去除 \n
                            topexsit = True
                    if topexsit == False :
                        if flag == 1:  # 插入表1数据
                            table1List1.append(row)
                        if flag == 2:  # 插入表1数据
                            table1List2.append(row)
                    if row[0].replace('\n', '').replace('\r', '') == "电厂名称":  # 下一页 没有表头 则继续上一页为同一个表格
                        flag = 0
                        flagtop = False
                    else:
                        flagtop = True
                else:
                    #sheet = workbook.add_sheet('Sheet1')  #添加sheet
                    # 1.正常读取行数 进行数据插入
                    if flag == 1:  # 插入表1数据
                        #sheet = workbook.add_sheet('Sheet1')  #添加sheet
                        #i = 0 # Excel起始位置
                        table1List1.append(row)
                        # for j in range(len(row)):#一行
                        #     sheet.write(i, j, row[j])#一列一列
                        # i += 1
                        # workbook.save(r'D:/AS3000/pdffile/1.xls')
                    if flag == 2:  # 插入表1数据
                        #sheet = workbook.add_sheet('Sheet1')  #添加sheet
                        #i = 0 # Excel起始位置
                        table1List2.append(row)
                        # for j in range(len(row)):#一行
                        #     sheet.write(i, j, row[j])#一列一列
                        # i += 1
                        # workbook.save(r'D:/AS3000/pdffile/2.xls')
                index = 1
                if (flag == 0 or flagtop) and len(table1) == len(top1) and ((table1 == np.array(top1)).all() or len(top1) == 0):  # 判定是否为表1的数据
                    flag = 1
                if (flag == 0 or flagtop) and len(table2) == len(top1) and ((table2 == np.array(top1)).all() or len(top1) == 0):
                    flag = 2

    b = 0
    print(table1List1)
    i = 1 # Excel起始位置
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet('Sheet1')  #添加sheet
    for a in range(len(table1.tolist())):
        sheet.write(0,a,table1.tolist()[a])
    for rows1 in table1List1:
        for j in range(len(rows1)):
            sheet.write(i, j, rows1[j])
        i += 1
    workbook.save(r'D:/AS3000/pdffile/'+now_time+str(b)+'.xls')
    print(table1List2)
    workbook = xlwt.Workbook()
    i = 1 # Excel起始位置
    sheet = workbook.add_sheet('Sheet1')  #添加sheet
    for a in range(len(table2.tolist())):
        sheet.write(0,a,table2.tolist()[a])
    for rows2 in table1List2:
        for j in range(len(rows2)):
            sheet.write(i, j, rows2[j])
        i += 1
    workbook.save(r'D:/AS3000/pdffile/'+now_time+str(b+1)+'.xls')

    # 2.特殊处理 同一张表在不同页上
    # 采用序号对比来进行判定
    # 3.特殊处理 一条序号横跨多行。 即多级电厂数据
