import pdfplumber
import numpy as np
import xlwt
from openpyxl import Workbook
from pdfplumber import table
#workbook = xlwt.Workbook()  #定义workbook
#sheet = workbook.add_sheet('Sheet1')  #添加sheet
with pdfplumber.open(r'D:\AS3000\PdfRead\20210304.pdf') as pdf:
    # 前提所有列必须固定
    table1 = np.array(['序号', '电厂名称', '考核费用净收入(万元)', '非停净收入(万元)', 'AGC补偿费用净收入(万元)', '旋转备用补偿费用净收入(万元)', '调峰补偿费用净收入(万元)'
                          , '无功补偿费用净收入(万元)', '黑启动补偿费用净收入(万元)', '冷备用补偿费用净收入(万元)', 'AVC补偿费用净收入(万元)', '风电调峰补偿费用净收入(万元)',
                       '返还缺额资金(万元)', '总净收入(万元)'])
    table2 = np.array(['序号', '电厂', '考核返还费用(万元)', '并网运行考核费用合计(万元)', '辅助服务考核费用(万元)', '考核费用净收入（万元）'])
    table3 = np.array(['序号', '电厂', '考核费用合计(万元)', '发电计划考核费用(万元)', '一次调频考核费用(万元)', 'AGC考核费用(万元)', '电压曲线合格率考核费用(万元)',
                       '调峰考核费用(万元)', '燃料管理考核费用(万元)', '技术监督考核费用(万元)', '安全管理考核费用(万元)', '调度管理考核费用(万元)', '检修管理考核费用(万元)',
                       '设备参数维护考核费用(万元)', '黑启动考核费用(万元)', '风电调度管理考核费用(万元)'])
    table4 = np.array(['序号', '电厂', '风电有功功率变化考核(万元)', '风电脱网考核(万元)', '风电发电计划考核费用(万元)', '风电风功率预测考核费用(万元)',
                       '风电低电压穿越能力考核(万元)', '风电动态无功补偿装置考核费用(万元)', '风电技术管理考核费用(万元)', '风电调度纪律考核费用(万元)', 'AVC考核费用(万元)'])
    table5 = np.array(['序号', '电厂', '机组', '机组容量(MW)', '等效非停时间(h)', '临时停运时间(h)', '非停总时间(h)', '考核标准时间(h)', '非停考核小时(h)',
                       '考核电量(MWh)', '非停考核费用(万元)', '非停返还费用(万元)', '非停净收入(万元)'])
    table6 = np.array(['序号', '电厂', '辅助服务补偿汇总(万元)', 'AGC补偿(万元)', '调峰补偿(万元)', '旋转备用补偿(万元)', '无功补偿(万元)', '黑启动补偿(万元)',
                       '风电调峰补偿(万元)', 'AVC补偿(万元)', '冷备用补偿(万元)'])

    table1List1 = []  # 表1数据集合
    table1List2 = []  # 表1数据集合
    table1List3 = []  # 表1数据集合
    table1List4 = []  # 表1数据集合
    table1List5 = []  # 表1数据集合
    table1List6 = []  # 表1数据集合
    #i = 0
    for page in pdf.pages:
        # page = pdf.pages[index]  # 页数 从0开始
        top1 = []
        index = 0
        for table in page.extract_tables():
            for row in table:

                if index == 0:
                    for cell in row:
                        top1.append("".join(cell.split()))  # 将第一行的集合去除 \n
                    if row[0] == "序号":  # 下一页 没有表头 则继续上一页为同一个表格
                        flag = 0
                else:
                    #sheet = workbook.add_sheet('Sheet1')  #添加sheet
                    # 1.正常读取行数 进行数据插入
                    if flag == 1:  # 插入表1数据
                        #sheet = workbook.add_sheet('Sheet1')  #添加sheet
                        #i = 0 # Excel起始位置
                        table1List1.append(row)
                        # for j in range(len(row)):#一行
                        #     sheet.write(i, j, row[j])#一列一列
                        # i += 1
                        # workbook.save(r'D:/AS3000/pdffile/1.xls')
                    if flag == 2:  # 插入表1数据
                        #sheet = workbook.add_sheet('Sheet1')  #添加sheet
                        #i = 0 # Excel起始位置
                        table1List2.append(row)
                        # for j in range(len(row)):#一行
                        #     sheet.write(i, j, row[j])#一列一列
                        # i += 1
                        # workbook.save(r'D:/AS3000/pdffile/2.xls')
                    if flag == 3:  # 插入表1数据
                        #sheet = workbook.add_sheet('Sheet1')  #添加sheet
                        #i = 0 # Excel起始位置
                        table1List3.append(row)
                        # for j in range(len(row)):#一行
                        #     sheet.write(i, j, row[j])#一列一列
                        # i += 1
                        # workbook.save(r'D:/AS3000/pdffile/3.xls')
                    if flag == 4:  # 插入表1数据
                        #sheet = workbook.add_sheet('Sheet1')  #添加sheet
                        #i = 0 # Excel起始位置
                        table1List4.append(row)
                        # for j in range(len(row)):#一行
                        #     sheet.write(i, j, row[j])#一列一列
                        # i += 1
                        # workbook.save(r'D:/AS3000/pdffile/4.xls')
                    if flag == 5:  # 插入表1数据
                        #sheet = workbook.add_sheet('Sheet1')  #添加sheet
                        #i = 0 # Excel起始位置
                        table1List5.append(row)
                        # for j in range(len(row)):#一行
                        #     sheet.write(i, j, row[j])#一列一列
                        # i += 1
                        # workbook.save(r'D:/AS3000/pdffile/5.xls')
                    if flag == 6:  # 插入表1数据
                        #sheet = workbook.add_sheet('Sheet1')  #添加sheet
                        #i = 0 # Excel起始位置
                        table1List6.append(row)
                        # for j in range(len(row)):#一行
                        #     sheet.write(i, j, row[j])#一列一列
                        # i += 1
                        # workbook.save(r'D:/AS3000/pdffile/6.xls')
                index = 1
                if flag == 0 and len(table1) == len(top1) and (table1 == np.array(top1)).all():  # 判定是否为表1的数据
                    flag = 1
                if flag == 0 and len(table2) == len(top1) and (table2 == np.array(top1)).all():
                    flag = 2
                if flag == 0 and len(table3) == len(top1) and (table3 == np.array(top1)).all():
                    flag = 3
                if flag == 0 and len(table4) == len(top1) and (table4 == np.array(top1)).all():
                    flag = 4
                if flag == 0 and len(table5) == len(top1) and (table5 == np.array(top1)).all():
                    flag = 5
                if flag == 0 and len(table6) == len(top1) and (table6 == np.array(top1)).all():
                    flag = 6

    print(table1List1)
    i = 1 # Excel起始位置
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet('Sheet1')  #添加sheet
    for a in range(len(table1.tolist())):
        sheet.write(0,a,table1.tolist()[a])
    for rows1 in table1List1:
        for j in range(len(rows1)):
            sheet.write(i, j, rows1[j])
        i += 1
    workbook.save(r'D:/AS3000/pdffile/1.xls')
    print(table1List2)
    workbook = xlwt.Workbook()
    i = 1 # Excel起始位置
    sheet = workbook.add_sheet('Sheet1')  #添加sheet
    for a in range(len(table2.tolist())):
        sheet.write(0,a,table2.tolist()[a])
    for rows2 in table1List2:
        for j in range(len(rows2)):
            sheet.write(i, j, rows2[j])
        i += 1
    workbook.save(r'D:/AS3000/pdffile/2.xls')
    print(table1List3)
    workbook = xlwt.Workbook()
    i = 1 # Excel起始位置
    sheet = workbook.add_sheet('Sheet1')  #添加sheet
    for a in range(len(table3.tolist())):
        sheet.write(0,a,table3.tolist()[a])
    for rows3 in table1List3:
        for j in range(len(rows3)):
            sheet.write(i, j, rows3[j])
        i += 1
    workbook.save(r'D:/AS3000/pdffile/3.xls')
    print(table1List4)
    workbook = xlwt.Workbook()
    i = 1 # Excel起始位置
    sheet = workbook.add_sheet('Sheet1')  #添加sheet
    for a in range(len(table4.tolist())):
        sheet.write(0,a,table4.tolist()[a])
    for rows4 in table1List4:
        for j in range(len(rows4)):
            sheet.write(i, j, rows4[j])
        i += 1
    workbook.save(r'D:/AS3000/pdffile/4.xls')
    print(table1List5)
    workbook = xlwt.Workbook()
    i = 1 # Excel起始位置
    sheet = workbook.add_sheet('Sheet1')  #添加sheet
    for a in range(len(table5.tolist())):
        sheet.write(0,a,table5.tolist()[a])
    for rows5 in table1List5:
        for j in range(len(rows5)):
            sheet.write(i, j, rows5[j])
        i += 1
    workbook.save(r'D:/AS3000/pdffile/5.xls')
    print(table1List6)
    workbook = xlwt.Workbook()
    i = 1 # Excel起始位置
    sheet = workbook.add_sheet('Sheet1')#添加sheet
    for a in range(len(table6.tolist())):
        sheet.write(0,a,table6.tolist()[a])
    for rows6 in table1List6:
        for j in range(len(rows6)):
            sheet.write(i, j, rows6[j])
        i += 1
    workbook.save(r'D:/AS3000/pdffile/6.xls')

    # 2.特殊处理 同一张表在不同页上
    # 采用序号对比来进行判定
    # 3.特殊处理 一条序号横跨多行。 即多级电厂数据
